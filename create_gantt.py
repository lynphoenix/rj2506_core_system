import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_gantt_chart():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "RJ2506 Project Schedule (Gantt)"

    # 定义颜色
    color_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    color_phase0 = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid') # 浅绿
    color_phase1 = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid') # 浅蓝
    color_phase2 = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid') # 浅橙

    font_header = Font(color='FFFFFF', bold=True, size=11)
    font_bold = Font(bold=True)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 数据定义: (阶段, 任务, 负责人, 开始周, 持续周)
    tasks = [
        # Phase 0 (W1-W2)
        ("Phase 0", "[0.1] 环境与连线 (ROS 2, Orin基础环境)", "实施实习生/测试工程师", 1, 1),
        ("Phase 0", "[0.2] 相机外参标定 (Eye-in-Hand)", "实施实习生/测试工程师", 1, 1),
        ("Phase 0", "[0.3] 物理探针制作与TCP标定", "硬件组/实施实习生", 2, 1),
        ("Phase 0", "[0.4] Ground Truth采集与验证", "实施实习生/测试工程师", 2, 1),

        # Phase 1 (W3-W8)
        ("Phase 1", "[模块5] C++ Daemon与软件时间同步", "算法组 (底层核心)", 3, 2),
        ("Phase 1", "[模块1] 双臂全身协调控制 (WBC)", "算法组 (运动学)", 3, 3),
        ("Phase 1", "[模块3] 双目特征记忆提取器", "算法组 (视觉)", 4, 3),
        ("Phase 1", "[模块4] 虚拟IBVS视觉伺服", "算法组 (控制)", 5, 3),
        ("Phase 1", "[模块6] 阻抗控制与触底检测", "算法组 (交互)", 6, 2),
        ("Phase 1", "[模块2] 顶层状态机与任务编排 (Brain)", "算法组 (架构)", 7, 2),

        # Phase 2 (W9-W12)
        ("Phase 2", "[2.1] 遥操作真机数据采集 (50条)", "实施实习生/算法组", 9, 1),
        ("Phase 2", "[2.1] IsaacLab Sim2Real合成数据扩充", "算法组 (深度学习)", 9, 2),
        ("Phase 2", "[2.2] DP/ACT 模型训练与TensorRT部署", "算法组 (深度学习)", 10, 2),
        ("Phase 2", "[2.3] 全链路实车打通与边缘Case调试", "全员", 11, 2),
    ]

    total_weeks = 12

    # 设置表头
    headers = ["阶段 (Phase)", "任务名称 (Task Name)", "负责人 (Owner)"] + [f"W{i}" for i in range(1, total_weeks + 1)]
    for col_idx, text in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = text
        cell.fill = color_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_thin

    # 填充数据和绘制甘特图色块
    for row_idx, (phase, task, owner, start_week, duration) in enumerate(tasks, 2):
        # 基础信息
        sheet.cell(row=row_idx, column=1).value = phase
        sheet.cell(row=row_idx, column=2).value = task
        sheet.cell(row=row_idx, column=3).value = owner

        # 根据阶段设置基础信息的样式
        for col in range(1, 4):
            cell = sheet.cell(row=row_idx, column=col)
            cell.alignment = align_left if col == 2 else align_center
            cell.border = border_thin
            if col == 1:
                cell.font = font_bold

        # 确定当前任务的颜色
        if phase == "Phase 0": fill_color = color_phase0
        elif phase == "Phase 1": fill_color = color_phase1
        else: fill_color = color_phase2

        # 填充甘特图周数色块
        for col_idx in range(4, 4 + total_weeks):
            current_week = col_idx - 3
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            # 如果当前周在任务的 (开始周, 开始周+持续周-1) 范围内，就涂色
            if start_week <= current_week < start_week + duration:
                cell.fill = fill_color

    # 调整列宽
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 45
    sheet.column_dimensions['C'].width = 25
    for i in range(4, 4 + total_weeks):
        sheet.column_dimensions[get_column_letter(i)].width = 5

    # 保存文件
    file_path = "docs/RJ2506_Project_Schedule_Gantt_V1.xlsx"
    wb.save(file_path)
    print(f"✅ Gantt chart generated at: {file_path}")

if __name__ == "__main__":
    create_gantt_chart()
