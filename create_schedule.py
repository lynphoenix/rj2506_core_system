import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Define the data for the project schedule
data = [
    # Phase 0
    ['Phase 0', '0.1 环境与连线', '系统刷机（不含实时补丁）、ROS 2 安装、外设连通性测试', 0.5, '实施组 (Intern/测试)', '无', 'ros2 --version 正常，能读取相机和底盘数据'],
    ['Phase 0', '0.2 相机外参标定 (Eye-in-Hand)', '使用 ChArUco 板，采集30个极限位姿，跑脚本解算 camera_to_flange 外参', 0.5, '实施组 (Intern/测试)', '0.1 环境与连线', '生成 rj2506_extrinsics.yaml 文件'],
    ['Phase 0', '0.3 绝对精度验证 (TCP 探针)', '制作物理探针，手工戳出 4 个验证点的绝对坐标，记录 Excel 供算法组验证', 0.5, '实施组 (Intern/测试)', '0.2 相机外参标定', '完成 4 点坐标采集表格，交给算法组'],
    ['Phase 0', '0.4 盲插定妆照采集 (金标准)', '人工将零件完美插入卡槽后抬高 15cm，拍摄 2D 目标金标准照片', 0.5, '实施组 (Intern/测试)', '无', '生成清晰的 golden_template.jpg'],
    ['Phase 0', '0.5 (算法组) 软件时间插值误差验证', '高速挥舞测试 Daemon RingBuffer 插值补偿，比对秒表与电机时间戳', 1.0, '算法组 (Algorithms)', '0.1 守护进程开发', '抖动误差 (Drift Variance) < 10ms'],
    ['Phase 0', '0.6 (算法组) 重投影误差验证', '利用相机内参与标定外参，计算物理坐标在图像上的虚拟投影与实际像素差', 1.0, '算法组 (Algorithms)', '0.2, 0.3', '重投影误差 RMSE < 1.5 像素'],
    
    # Phase 1
    ['Phase 1', '1.1 C++ 守护进程与软件时间同步', '在 1000Hz 纯净线程管理 CANopen，利用环形缓冲区解决因果混淆', 2.0, '算法组 (Algorithms)', '无', 'Daemon 稳定运行，无丢帧，时间戳误差<10ms'],
    ['Phase 1', '1.2 虚拟 IBVS 视觉伺服', '推导包含深度的 6x6 图像雅可比矩阵，防倾斜发散，伴随矩阵速度映射', 2.0, '算法组 (Algorithms)', '0.6 重投影验证通过', '虚拟相机可基于金标准输出 $\Delta v$'],
    ['Phase 1', '1.3 双目特征提取器 (3D 记忆)', '双轨特征提取（几何算法 + SuperPoint/LightGlue），稳定提取 3D 锚点', 1.5, '算法组 (Algorithms)', '1.2 IBVS 数学推导', '恶劣光照下稳定提取 N >= 4 个 3D 特征点'],
    ['Phase 1', '1.4 双臂全身协调控制 (WBC)', '计算闭链运动学，解算双臂干涉避障与双手抓取轨迹规划', 2.5, '算法组 (Algorithms)', '无', '双臂无碰撞到达目标点，平稳抱持料框'],
    ['Phase 1', '1.5 顶层状态机与大脑节点', '设计 Behavior Tree 管理算法生命周期，处理边缘情况与异常恢复', 1.5, '算法组 (Algorithms)', '1.1, 1.2, 1.3, 1.4', '完整串联 Nav2 -> WBC -> IBVS，模拟异常可恢复'],
    ['Phase 1', '1.6 力控、触底检测与物理顺应', '在 CANopen 外层封装阻抗控制，依靠 $\Delta I / \Delta t$ 识别触底并释放', 1.5, '算法组 (Algorithms)', '1.1 守护进程', '安全感知碰撞，触底后双夹爪同步释放'],
    
    # Phase 2
    ['Phase 2', '2.1 遥操作数据采集 (Leader-Follower)', '录制 50 条专家轨迹，利用 Daemon 严格对齐图像与关节状态', 1.5, '实施组 & 算法组', 'Phase 0, 1.1', '获取 100 条高质量 HDF5 数据集 (因果误差<10ms)'],
    ['Phase 2', '2.2 IsaacLab 仿真与数据增强', '将真机数据导入 IsaacLab，域随机化生成 10000 条合成数据', 1.5, '算法组 (Algorithms)', '2.1 数据采集', '成功扩充包含光照、纹理变异的合成数据集'],
    ['Phase 2', '2.3 ACT 模型训练 (DP)', '在 H100 训练模型，预测未来 k 步 14 轴关节空间目标轨迹 ($q$)', 2.0, '算法组 (Algorithms)', '2.2 数据增强', '模型收敛，验证集动作误差达标'],
    ['Phase 2', '2.4 TensorRT 边缘端部署', '导出 ONNX 并在 Jetson Orin 编译 FP16 引擎', 1.0, '算法组 (Algorithms)', '2.3 模型训练', 'Orin 端 ACT 推理延迟 < 30ms'],
    ['Phase 2', '2.5 大脑节点实车全链路联调', 'Brain Node 在真机串联 8 段管线，执行异常重试与安全降级', 2.0, '全体 (All Teams)', 'Phase 1, 2.4', '连续 20 次盲插搬运成功，0 人工接管']
]

df = pd.DataFrame(data, columns=['阶段 (Phase)', '任务名称 (Task Name)', '任务描述 (Description)', '工期 (Duration/Weeks)', '执行角色 (Assigned Role)', '前置依赖 (Dependencies)', '验收标准 (Acceptance Criteria)'])

# Calculate Phase Durations (approximate, max of parallel tracks)
# Phase 0: ~2 weeks total
# Phase 1: ~6 weeks total
# Phase 2: ~4 weeks total

# Create Excel with formatting
excel_path = 'docs/RJ2506_Project_Schedule_V1.xlsx'
writer = pd.ExcelWriter(excel_path, engine='openpyxl')
df.to_excel(writer, sheet_name='Project Schedule', index=False)

workbook = writer.book
worksheet = writer.sheets['Project Schedule']

# Formatting Styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
border_style = Side(border_style="thin", color="000000")
border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

# Apply Header formatting
for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Apply data formatting
for row in worksheet.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=7):
    # Phase color coding
    phase = row[0].value
    if phase == 'Phase 0':
        row_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    elif phase == 'Phase 1':
        row_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    elif phase == 'Phase 2':
        row_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
    for cell in row:
        cell.fill = row_fill
        cell.alignment = alignment
        cell.border = border

# Set column widths
worksheet.column_dimensions['A'].width = 12
worksheet.column_dimensions['B'].width = 35
worksheet.column_dimensions['C'].width = 50
worksheet.column_dimensions['D'].width = 15
worksheet.column_dimensions['E'].width = 20
worksheet.column_dimensions['F'].width = 25
worksheet.column_dimensions['G'].width = 45

writer.close()
print(f"Project Schedule successfully created at {excel_path}")
